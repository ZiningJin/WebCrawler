from flask import Blueprint, request, jsonify, Response, current_app
from werkzeug.utils import secure_filename
import os
import io
import re
import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from models import db, Bci_uniqueid_tbl
from utils.bci_utils import bci_util_concat
from utils.pec_utils import pec_util_login, pec_util_auth_code, pec_util_criteria, pec_util_extract, pec_util_email, pec_util_transform
from utils.cceup_utils import patterns_abcd, patterns_e, cceup_util_extract_abcd, cceup_util_extract_e, patterns_co, patterns_person, patterns_phone, patterns_city, cceup_extract_info, cceup_clean_text, co_cases, cceup_clean_co

api = Blueprint('api', __name__)

# --------- BCI Step1 File Upload ---------
@api.route('/bci_upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return jsonify({'message': 'No file part'}), 400
    files = request.files.getlist('files')

    for file in files:
        if file.filename == '':
            return jsonify({'message': 'No selected file'}), 400
        if file:
            filename = secure_filename(file.filename)
            # the files will be saved at server/bci_data/
            file_path = os.path.join('bci_data', filename)
            file.save(file_path)
    return jsonify({'message': 'Files successfilly uploaded'})

# --------- BCI Step2 Data Concatenation and Processing ---------
@api.route('/bci_concat', methods=['GET'])
def bci_concat():
    # check the current working directory for os.walk 
    # print("Current Working Directory:",os.getcwd())
    dfs = pd.DataFrame()
    expected_columns = ['Role on Project', 'Contact First Name', 'Contact Surname', 'Company Phone', 'Company Email']
    # concatenate all files under server/bci_data into 1 dfs
    for dirname, _, filenames in os.walk('bci_data'):
        for filename in filenames:
            if filename.endswith('.xlsx'):
                file_path = os.path.join(dirname, filename)
                try:
                    df = pd.read_excel(file_path)
                    # Check if the DataFrame contains the expected columns
                    if all(column in df.columns for column in expected_columns):
                        dfs = pd.concat([dfs, df])
                    else:
                        print(f"File {filename} skipped: missing required columns")
                except Exception as e:
                    print(f"Error reading file {filename}: {e}")
    
    if dfs.empty:
        return jsonify({'error': 'No valid data found in uploaded files'})

    dfs = dfs.reset_index(drop=True)

    # preserve the 1st row of field Owner Enterprice
    dfs['Owner Enterprise'] = dfs['Role on Project'] + ':' + dfs['Contact First Name'] + ' ' + dfs['Contact Surname'] + ',' + dfs['Company Phone'] + ',' + dfs['Company Email']
    dfs_temp = dfs.groupby(by='Project ID').apply(lambda row: bci_util_concat(row)).reset_index()

    dfs_first = dfs.groupby(by=['Project ID'])[['Project Type', 'Project Name', 'Project Stage',
           'Project Status', 'Project Address', 'Project Country', 'Company Name',
           'Company Phone', 'Company Email', 'Contact First Name', 'Contact Surname']].first()
    dfs_first.reset_index(inplace=True)

    dfs_concat = pd.merge(left=dfs_first, right=dfs_temp, how='inner', left_on='Project ID', right_on='Project ID')
    # save dfs_concat for Step3
    dfs_concat.to_excel('bci_data/dfs_concat.xlsx', index=False)
    # send the row counts to Frontend
    row_count = dfs_concat.shape[0]
    return jsonify({'row_count': row_count, 'message': 'Data concated'})   

# --------- BCI Step3&4 Drop Duplicate & Update Database ---------  
@api.route('/bci_dedup_update', methods=['GET'])
def bci_dedup_update():
    # read dfs_concat from server/bci_data
    dfs_concat = pd.read_excel('bci_data/dfs_concat.xlsx')
    # connect to db and fetch existing projectids
    existing_ids = pd.read_sql('SELECT projectid FROM leads.bci_uniqueid_tbl', con=db.engine)
    # drop duplicates
    dfs_unique = dfs_concat[~dfs_concat['Project ID'].isin(existing_ids['projectid'])]
    dfs_unique = dfs_unique.reset_index(drop=True)
    # update db with new unique projectids
    new_unique_projects = dfs_unique[['Project ID', 'Project Country']].drop_duplicates()
    for idx, row in new_unique_projects.iterrows():
        new_entry = Bci_uniqueid_tbl(projectid=row['Project ID'], country=row['Project Country'])
        db.session.add(new_entry)
    db.session.commit()
    # save dfs_unique to server/bci_data
    dfs_unique.to_excel('bci_data/dfs_unique.xlsx', index=False)
    # return row count of dfs_unique to Frontend
    row_count = dfs_unique.shape[0]
    return jsonify({'row_count': row_count, 'message': 'Data deduplicated and database updated'})

# --------- BCI Step7 Download Archived ProjectID and Country ---------  
@api.route('/bci_download_archive', methods=['GET'])
def bci_download_archive():
    # query the database
    data = Bci_uniqueid_tbl.query.with_entities(Bci_uniqueid_tbl.projectid, Bci_uniqueid_tbl.country).all()
    # convert to df
    df = pd.DataFrame(data, columns=['projectid', 'country'])
    # convert df to csv, io.BytesIO to create in-memory buffer instead of io.StringIO of in-memory text
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    # rewinds the buffer
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=bci_uniqueids_tbl_archive.csv"})

# --------- BCI Step5 Data Transformation ---------
@api.route('/bci_transform', methods=['GET'])
def bci_transform():
    # read dfs_unique
    dfs_unique = pd.read_excel('bci_data/dfs_unique.xlsx')
    # rename columns
    dfs_unique = dfs_unique.rename(columns={
        'Project ID':'External ID','Project Country':'Country','Company Name':'Company',
        'Project Type':'Project Description','Project Stage': 'Project Stage 1','Company Phone': 'Phone',
        'Company Email': 'Email Address'
        })
    dfs_unique['Last Name'] = dfs_unique['Contact First Name']+dfs_unique['Contact Surname']
    dfs_unique['Last Name'] = dfs_unique['Last Name'].replace('--', 'Undisclosed')
    dfs_unique['Phone'] = dfs_unique['Phone'].apply(lambda x: x.split('/')[0] if '/' in x else x)
    dfs_unique['Country'] = dfs_unique['Country'].apply(lambda x: 'Viet Nam' if x=='Vietnam' else x)
    # the nulls and the fixed
    dfs_unique['WG_AMT'] = 'WG'
    dfs_unique['Brands'] = 'Ingersoll Rand'
    dfs_unique['Product Category'] = 'Compressor'
    dfs_unique['Lead Source'] = 'Project Database'
    dfs_unique['Lead Source 1'] = 'BCI'

    dfs_unique['Marketing_ID']= "" 
    dfs_unique['State']= "" 
    dfs_unique['City']= "" 
    dfs_unique['Industry']= "" 
    dfs_unique['Lead Source 2']= "" 
    dfs_unique['Lead Source 3']= "" 
    dfs_unique['Lead Source 4']= "" 
    dfs_unique['Start Time']= "" 
    dfs_unique['Project Property']= "" 
    dfs_unique['Design Enterprise']= "" 
    dfs_unique['EPC Enterprise']= "" 
    # sort
    dfs_unique = dfs_unique[['Project Name','External ID','Marketing_ID', 
        'Last Name', 'Phone', 'Email Address','WG_AMT', 'Brands', 'Product Category','Country',
        'State', 'City', 'Industry','Lead Source','Lead Source 1','Lead Source 2', 'Lead Source 3', 'Lead Source 4','Start Time',
        'Company','Project Description','Project Stage 1', 'Project Property','Project Address',
        'Owner Enterprise','Design Enterprise','EPC Enterprise']]
    # convert df to csv, io.BytesIO to create in-memory buffer instead of io.StringIO of in-memory text
    output = io.BytesIO()
    dfs_unique.to_csv(output, index=False, encoding='utf-8')
    # rewinds the buffer
    output.seek(0)

    # before return, del all files under folder server/bci_data
    for filename in os.listdir('bci_data'):
        file_path = os.path.join('bci_data', filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s'%(file_path,e)) 

    return Response(output, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=bci_dfs_unique.csv"})


# --------- PEC Step1 Login  ---------
browser_sessions = {}

def generate_unique_session_id():
    return str(uuid.uuid4())

@api.route('/pec_login', methods=['POST'])
def pec_login():
    data = request.json
    url = "https://www.industrialinfo.com/dash/project_search.jsp"
    browser = pec_util_login(url, data['username'], data['password'])
    if browser:
        session_id = generate_unique_session_id()
        browser_sessions[session_id] = browser
        # Test the session_id from pec_login with that from pec_process
        current_app.logger.info(f"Session created: {session_id}")
        return jsonify({"message": "Login successful", "session_id": session_id}), 200
    else:
        current_app.logger.error("Login failed")
        return jsonify({"error": "Login failed"}), 400

# --------- PEC Process  ---------
@api.route('/pec_process', methods=['POST'])
def pec_process():
    data = request.json
    session_id = data['session_id']
    # Test the session_id from pec_login with that from pec_process
    current_app.logger.info(f"Attempting to use session: {session_id}")
    browser = browser_sessions.get(session_id)

    if not browser:
        return jsonify({"error": "Session not found or expired"}), 404

    browser = pec_util_auth_code(browser, data['auth_code'])
    if not browser:
        browser.quit()
        del browser_sessions[session_id]
        return jsonify({"error": "Authentication Failed"}), 400

    browser = pec_util_criteria(browser, data['start_date_str'], data['end_date_str'])
    projectINFOs = pec_util_extract(browser)
    if not projectINFOs:
        browser.quit()
        return jsonify({"error": "Extraction Failed"}), 400

    df = pec_util_transform(projectINFOs)
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    browser.quit()

    # Clean up the session
    del browser_sessions[session_id]

    # Clean up any previous data
    for filename in os.listdir('pec_data'):
        file_path = os.path.join('pec_data', filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s'%(file_path,e)) 
    # ...

    return Response(output, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=pec_data.csv"})


# --------- CCEUP Step1 Upload  ---------
@api.route('/cceup_upload', methods=['POST'])
def cceup_upload_file():
    # check if the file part is present in the request
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400
    file = request.files['file']
    # check if a file is actually selected
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400
    # Validate the file is an Excel file by checking its extension
    if file and file.filename.endswith(('.xls', '.xlsx')):
        filename = secure_filename(file.filename)
        # save the file in server/cceup_data
        file_path = os.path.join('cceup_data', filename)
        file.save(file_path)
        return jsonify({'message': 'File successfully uploaded'}), 200
    else:
        return jsonify({'message': 'Invalid file type, please upload an Excel file'}), 400

# --------- CCEUP Step2 Transformation  ---------
@api.route('/cceup_transform', methods=['GET'])
def cceup_transform():
    # read excel into df
    for dirname, _, filenames in os.walk('cceup_data'):
        for filename in filenames:
            if filename.endswith('.xlsx'):
                file_path = os.path.join(dirname, filename)
                try:
                    df = pd.read_excel(file_path)
                except Exception as e:
                    print(f"Error reading file {filename}: {e}")
    
    if df.empty:
        return jsonify({'error': 'No valid data found in uploaded file'})
    # extract ABCD
    df = cceup_util_extract_abcd(df, '内容',patterns_abcd)
    df['ABD'] = df['A'] + df['B'] + df['D']
    # extract E
    df = cceup_util_extract_e(df, '内容', patterns_e)
    # extract E_block1
    df['E_block1'] = df['E'].apply(lambda x: re.split('\n\n',x)[0])
    # extract company
    df['Contact_co'] = cceup_extract_info(df, 'E', patterns_co)
    # extract person
    df['Contact_person'] = cceup_extract_info(df, 'E', patterns_person)
    # extract phone
    df['Contact_phone'] = cceup_extract_info(df, 'E', patterns_phone)
    # clean person
    df = cceup_clean_text(df, 'Contact_person','[^\u4e00-\u9fa5]+')
    # clean phone
    df = cceup_clean_text(df, 'Contact_phone', '[^0-9\-]+')
    # clean company
    df['Contact_co'] = df['Contact_co'].apply(cceup_clean_co)
    # extract city
    df['City'] = cceup_extract_info(df, 'C', patterns_city, default='').apply(lambda x: x + '市' if x else x)
    # clean E_block1
    df['E_block1'] = df['E'].apply(lambda x: re.split('\n\n',x)[0] if re.split('\n\n',x)[0]!='' else re.split('\n\n',x)[1])
    # save df to server/cceup_data
    df.to_excel('cceup_data/df_cceup.xlsx', index=False)
    # return row count of dfs_unique to Frontend
    row_count = df.shape[0]
    return jsonify({'row_count': row_count, 'message': 'Data Transformed Successfully'})

# --------- CCEUP Step3 Download  ---------
@api.route('/cceup_download', methods=['GET'])
def cceup_download():
    # Read df_cceup.xlsx
    df = pd.read_excel('cceup_data/df_cceup.xlsx')

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Apply formatting
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=11)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                ws.row_dimensions[r_idx].height = 14
            else:
                cell.font = cell_font
            ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].auto_size = True
    
    # Apply filters to all columns
    ws.auto_filter.ref = ws.dimensions

    # Save workbook to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Delete all files under folder 'cceup_data'
    for filename in os.listdir('cceup_data'):
        file_path = os.path.join('cceup_data', filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f'Failed to delete {file_path}. Reason: {e}')

    # Return the Excel file to download
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment;filename=df_cceup.xlsx"})
